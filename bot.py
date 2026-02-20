import asyncio
import sqlite3
from datetime import datetime
from aiogram import Bot, Dispatcher
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton
from aiogram.filters import Command
from openpyxl import load_workbook

# ================= НАСТРОЙКИ =================

import os
TOKEN = os.getenv("7380384121:AAESdUc0HQAjDLqBVLw5deYOFOyO5Fbmyro")
GROUP_ID = -5213831642

bot = Bot(token=TOKEN)
dp = Dispatcher()

user_carts = {}
waiting_for_phone = {}

# ================= БАЗА ДАННЫХ =================

conn = sqlite3.connect("orders.db")
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS orders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    date TEXT,
    client_name TEXT,
    username TEXT,
    phone TEXT,
    items TEXT,
    total INTEGER
)
""")

conn.commit()

# ================= EXCEL =================

def load_products_from_excel():
    workbook = load_workbook("products.xlsx")
    sheet = workbook.active

    products = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        category, name, price = row

        if category not in products:
            products[category] = []

        products[category].append({
            "name": name,
            "price": price
        })

    return products


products = {}

def reload_products():
    global products
    products = load_products_from_excel()

reload_products()

# ================= КЛАВИАТУРЫ =================

def get_main_keyboard():
    buttons = [[KeyboardButton(text=category)] for category in products.keys()]
    buttons.append([KeyboardButton(text="🛒 Корзина")])
    buttons.append([KeyboardButton(text="🔄 Обновить прайс")])
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)


def get_products_keyboard(category):
    buttons = []
    for product in products[category]:
        text = f"{product['name']} - {product['price']} ₽"
        buttons.append([KeyboardButton(text=text)])

    buttons.append([KeyboardButton(text="⬅️ Назад")])
    buttons.append([KeyboardButton(text="🛒 Корзина")])

    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)

# ================= СТАРТ =================

@dp.message(Command("start"))
async def start_handler(message: Message):
    user_carts[message.from_user.id] = []
    await message.answer(
        "🥩 Добро пожаловать!\n\nВыберите категорию:",
        reply_markup=get_main_keyboard()
    )

# ================= ОБРАБОТКА =================

@dp.message()
async def message_handler(message: Message):
    user_id = message.from_user.id
    text = message.text

    # Обновить прайс
    if text == "🔄 Обновить прайс":
        reload_products()
        await message.answer("✅ Прайс обновлён!", reply_markup=get_main_keyboard())
        return

    # Назад
    if text == "⬅️ Назад":
        await message.answer("Выберите категорию:", reply_markup=get_main_keyboard())
        return

    # Корзина
    if text == "🛒 Корзина":
        cart = user_carts.get(user_id, [])

        if not cart:
            await message.answer("Корзина пуста 🛒")
            return

        total = sum(item["price"] * item["quantity"] for item in cart)

        cart_text = "🛒 Ваша корзина:\n\n"
        for item in cart:
            subtotal = item["price"] * item["quantity"]
            cart_text += f"{item['name']} x{item['quantity']} = {subtotal} ₽\n"

        cart_text += f"\n💰 Итого: {total} ₽\n\n"
        cart_text += "Чтобы удалить товар — нажмите кнопку ниже."

        buttons = []
        for item in cart:
            buttons.append([KeyboardButton(text=f"❌ {item['name']}")])

        buttons.append([KeyboardButton(text="🗑 Очистить корзину")])
        buttons.append([KeyboardButton(text="✅ Оформить заказ")])
        buttons.append([KeyboardButton(text="⬅️ Назад")])

        await message.answer(
            cart_text,
            reply_markup=ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)
        )
        return

    # Удаление товара
    if text.startswith("❌ "):
        product_name = text.replace("❌ ", "")
        cart = user_carts.get(user_id, [])

        for item in cart:
            if item["name"] == product_name:
                cart.remove(item)
                await message.answer(f"❌ {product_name} удалён из корзины.")
                return
        return

    # Очистить корзину
    if text == "🗑 Очистить корзину":
        user_carts[user_id] = []
        await message.answer("🗑 Корзина очищена.")
        return

    # Оформление заказа
    if text == "✅ Оформить заказ":
        waiting_for_phone[user_id] = True
        await message.answer("📱 Введите ваш номер телефона:")
        return

    # Получаем телефон
    if waiting_for_phone.get(user_id):
        phone = text
        cart = user_carts.get(user_id, [])

        if not cart:
            await message.answer("Корзина пуста.")
            return

        total = sum(item["price"] * item["quantity"] for item in cart)

        order_text = (
            "🔥 НОВЫЙ ЗАКАЗ\n\n"
            f"👤 {message.from_user.full_name}\n"
            f"📱 {phone}\n"
            f"🆔 @{message.from_user.username}\n\n"
        )

        items_text = ""
        for item in cart:
            subtotal = item["price"] * item["quantity"]
            order_text += f"{item['name']} x{item['quantity']} = {subtotal} ₽\n"
            items_text += f"{item['name']} x{item['quantity']}, "

        order_text += f"\n💰 Итого: {total} ₽"

        # Сохраняем в базу
        cursor.execute("""
        INSERT INTO orders (date, client_name, username, phone, items, total)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            message.from_user.full_name,
            message.from_user.username,
            phone,
            items_text,
            total
        ))

        conn.commit()

        # Отправляем в группу
        await bot.send_message(GROUP_ID, order_text)

        await message.answer("✅ Заказ отправлен! С вами свяжется менеджер.")

        user_carts[user_id] = []
        waiting_for_phone[user_id] = False
        return

    # Выбор категории
    if text in products:
        await message.answer(
            f"Вы выбрали {text}\n\nВыберите товар:",
            reply_markup=get_products_keyboard(text)
        )
        return

    # Добавление товара
    for category, items in products.items():
        for item in items:
            expected = f"{item['name']} - {item['price']} ₽"
            if text == expected:
                cart = user_carts.setdefault(user_id, [])

                for cart_item in cart:
                    if cart_item["name"] == item["name"]:
                        cart_item["quantity"] += 1
                        break
                else:
                    cart.append({
                        "name": item["name"],
                        "price": item["price"],
                        "quantity": 1
                    })

                await message.answer(f"✅ {item['name']} добавлен в корзину!")
                return

    await message.answer("Пожалуйста, используйте кнопки меню.")

# ================= ЗАПУСК =================

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())